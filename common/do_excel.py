# -*- coding:utf-8 _*-
""" 

@function： 
"""
import json
import os

import openpyxl

from common import contants


class Case:


    def __init__(self):
        self.id = None
        self.url = None
        self.data = None
        self.title = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None


class DoExcel:
    file_name = None

    def __init__(self, file_name):
        try:
            self.file_name = file_name
            self.workbook = openpyxl.load_workbook(filename=file_name)
        except FileNotFoundError as e:
            print('{0} not found, please check file path'.format(file_name))
            raise e

    def get_cases(self, sheet_name):
        sheet_name = sheet_name
        sheet = self.workbook[sheet_name]
        max_row = sheet.max_row
        cases = []
        for r in range(2, max_row + 1):
            case = Case()
            case.id = sheet.cell(row=r, column=1).value
            case.title = sheet.cell(row=r, column=2).value
            case.url = sheet.cell(row=r, column=3).value
            case.data = sheet.cell(row=r, column=4).value
            case.method = sheet.cell(row=r, column=5).value
            case.expected = sheet.cell(row=r, column=6).value
            if type(case.expected) == int:
                case.expected = str(case.expected)
            cases.append(case)

        return cases

    def write_result(self, sheet_name, row, actual, result):
        sheet = self.workbook[sheet_name]
        sheet.cell(row, 7).value = actual
        sheet.cell(row, 8).value = result
        self.workbook.save(filename=self.file_name)


if __name__ == '__main__':
   do_excel = DoExcel('E:\ZQ_code\datas\cases.xlsx')
   cases = do_excel.get_cases('继承')
   print(cases[1].id)